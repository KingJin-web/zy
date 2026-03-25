import json
import os
import requests
import time
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill



def print_disclaimer():
    """打印免责声明"""
    print("=" * 60)
    print("【免责声明】")
    print("1. 本程序仅为技术学习和个人账号信息查询使用，请勿用于商业用途或非法操作。")
    print("2. 使用本程序需确保您对所查询的快手账号拥有合法所有权或使用权，严禁查询他人账号信息。")
    print("3. 快手平台API可能随时变更，导致程序无法正常运行，开发者不保证程序的稳定性和可用性。")
    print("4. 任何因使用本程序违反快手平台用户协议、法律法规或侵犯他人权益的行为，均由使用者自行承担全部责任。")
    print("5. 开发者对使用本程序产生的任何直接或间接损失不承担任何法律责任。")
    print("6. 本程序为临时查询工具，建议您在使用完毕后24小时内删除程序及相关账号配置信息，以保障账号安全。")
    print("确认了解并同意以上条款后，程序将继续运行...")
    print("快手极速版以及普通版本cookie有的账号是通用的普通版(kpn=KUAISHOU) 极速版(kpn=NEBULA)更改cookie的kpn即可")
    print("\n请设置环境变量 ksck，格式为:")
    print("备注#cookie#salt#ip|端口|用户名|密码|到期日期")
    print("=" * 60 + "\n")


# 比率映射关系
RATIO_MAP = {
    19: 0.00009,
    28: 0.00008,
    37: 0.00007,
    46: 0.00006,
    55: 0.00005,
    64: 0.00004,
    73: 0.00003,
    82: 0.00002,
    91: 0.00001
}

# 读取环境变量中的参数，默认46
SELECTED_PARAM = int(os.getenv("RATIO_PARAM", 46))


class KuaishouClient:
    def __init__(self, index, cookie, remark=""):
        self.index = index
        self.remark = remark
        self.cookie = cookie
        self.user_name = "未知用户"

        # 余额信息
        self.coin_balance = 0
        self.cash_balance = 0
        self.total_balance = 0

        # 今日和昨日金币
        self.today_coin = 0
        self.yesterday_coin = 0

        # 兑换后金额
        self.today_coin_cash = 0
        self.yesterday_coin_cash = 0

        # 流水记录
        self.coin_records = []
        self.cash_records = []

        # 基础配置
        self.headers = {
            'User-Agent': "Mozilla/5.0 (Linux; Android 16; PTP-AN00 Build/HONORPTP-AN00; wv) AppleWebKit/537.36 (KHTML, like Gecko) Version/4.0 Chrome/139.0.7258.158 Mobile Safari/537.36 Yoda/3.2.17-rc2 Kwai/13.9.30.44872 OS_PRO_BIT/64 MAX_PHY_MEM/15272 KDT/PHONE AZPREFIX/az1 ICFO/0 StatusHT/38 TitleHT/44 NetType/NR ISLP/0 ISDM/0 ISLB/0 locale/zh-cn SHP/2665 SWP/1264 SD/3.5 CT/0 ISLM/0",
            'Accept-Encoding': "gzip, deflate, br, zstd",
            'sec-ch-ua-platform': "\"Android\"",
            'sec-ch-ua': "\"Not;A=Brand\";v=\"99\", \"Android WebView\";v=\"139\", \"Chromium\";v=\"139\"",
            'sec-ch-ua-mobile': "?1",
            'X-Requested-With': "com.smile.gifmaker",
            'Sec-Fetch-Site': "same-origin",
            'Sec-Fetch-Mode': "cors",
            'Sec-Fetch-Dest': "empty",
            'Referer': "https://encourage.kuaishou.com/kwai/profit?tab=coin&layoutType=4",
            'Accept-Language': "zh-CN,zh;q=0.9,en-US;q=0.8,en;q=0.7",
            'Cookie': self.cookie
        }

    def send_request(self, url, method="get", params=None):
        """发送请求"""
        try:
            if method == "get":
                response = requests.get(url, headers=self.headers, params=params, timeout=10)
            else:
                response = requests.post(url, headers=self.headers, data=params, timeout=10)

            if response.status_code == 200:
                return response.json()
            else:
                print(f"请求失败，状态码: {response.status_code}")
                return None
        except Exception as e:
            print(f"请求异常: {e}")
            return None

    def query_withdraw_info(self):
        """查询提现信息（包含用户名和余额）"""
        url = "https://encourage.kuaishou.com/rest/wd/encourage/account/withdraw/info"
        params = {
            "source": "normal",
            "imei": "",
            "oaid": "279b0911-a941-4e80-b5fd-91f1c1ea2644",
            "idfa": ""
        }

        result = self.send_request(url, params=params)
        if result and result.get("result") == 1:
            data = result.get("data", {})

            # 获取用户名
            self.user_name = data.get("nickname", "未知用户")

            # 获取账户信息
            account = data.get("account", {})
            self.coin_balance = account.get("coinAmount", 0)
            self.cash_balance = account.get("cashAmount", 0)  # 单位：分
            self.total_balance = account.get("accumulativeAmount", 0)  # 单位：分

            return True
        return False

    def query_cash_records(self):
        """查询现金流水记录"""
        url = "https://encourage.kuaishou.com/rest/wd/encourage/account/detail"
        params = {
            "__NS_sig3": "6272350512f90404a43e5b3d3a3b5b3320a219ee1616ba201b5f2d2d2b2b28291636",
            "sigCatVer": "1",
            "accountType": "cash",
            "cursor": ""
        }

        result = self.send_request(url, params=params)
        if result and result.get("result") == 1:
            data = result.get("data", {})
            records = data.get("datas", [])

            for record in records[:3]:
                self.cash_records.append({
                    'time': self._format_timestamp(record.get("createTime", 0)),
                    'title': record.get("title", ""),
                    'amount': record.get("displayAmount", "0"),
                    'direction': record.get("direction", "")
                })
            return True
        return False

    def query_coin_records(self):
        """查询金币流水记录"""
        url = "https://encourage.kuaishou.com/rest/wd/encourage/account/detail"
        params = {
            "__NS_sig3": "7262251502e91414b42e4a2d2a2b6006895f03fe0606aa3024c33d3d3b3b38390626",
            "sigCatVer": "1",
            "accountType": "coin",
            "cursor": ""
        }

        result = self.send_request(url, params=params)
        if result and result.get("result") == 1:
            data = result.get("data", {})
            records = data.get("datas", [])

            for record in records[:3]:
                self.coin_records.append({
                    'time': self._format_timestamp(record.get("createTime", 0)),
                    'title': record.get("title", ""),
                    'amount': record.get("displayAmount", "0"),
                    'direction': record.get("direction", "")
                })
            return True
        return False

    def _format_timestamp(self, timestamp):
        """格式化时间戳"""
        if timestamp:
            try:
                dt = datetime.fromtimestamp(timestamp / 1000)
                return dt.strftime("%Y-%m-%d %H:%M")
            except:
                return "未知时间"
        return "未知时间"

    def calculate_coin_stats(self):
        """计算今日和昨日金币收入及兑换后金额"""
        today = datetime.now().date()
        yesterday = today - timedelta(days=1)

        self.today_coin = 0
        self.yesterday_coin = 0

        url = "https://encourage.kuaishou.com/rest/wd/encourage/account/detail"
        params = {
            "__NS_sig3": "7262251502e91414b42e4a2d2a2b6006895f03fe0606aa3024c33d3d3b3b38390626",
            "sigCatVer": "1",
            "accountType": "coin",
            "cursor": ""
        }

        result = self.send_request(url, params=params)
        if result and result.get("result") == 1:
            data = result.get("data", {})
            records = data.get("datas", [])

            for record in records:
                if record.get("direction") != "IN":
                    continue

                amount = record.get("amount", 0)
                create_time = record.get("createTime", 0)

                if create_time:
                    record_date = datetime.fromtimestamp(create_time / 1000).date()

                    if record_date == today:
                        self.today_coin += amount
                    elif record_date == yesterday:
                        self.yesterday_coin += amount

        # 计算兑换后金额
        ratio = RATIO_MAP.get(SELECTED_PARAM, 0.00006)  # 默认46对应的比率
        self.today_coin_cash = round(self.today_coin * ratio, 2)
        self.yesterday_coin_cash = round(self.yesterday_coin * ratio, 2)

        return True

    def query_all_data(self):
        """查询所有数据"""
        print(f"正在查询账号[{self.index}] {self.remark}...")

        # 查询提现信息（包含用户名和余额）
        if not self.query_withdraw_info():
            print(f"账号[{self.index}] {self.remark} 提现信息查询失败")
            return False

        # 查询现金流水
        if not self.query_cash_records():
            print(f"账号[{self.index}] {self.remark} 现金流水查询失败")

        # 查询金币流水
        if not self.query_coin_records():
            print(f"账号[{self.index}] {self.remark} 金币流水查询失败")

        # 计算金币统计
        self.calculate_coin_stats()

        return True

    def get_display_data(self):
        """获取显示数据"""
        # 现金余额从分转换为元
        cash_yuan = self.cash_balance / 100 if self.cash_balance else 0
        total_yuan = self.total_balance / 100 if self.total_balance else 0

        return {
            "index": self.index,
            "user_name": self.user_name,
            "remark": self.remark,
            "today_coin": self.today_coin,
            "yesterday_coin": self.yesterday_coin,
            "today_coin_cash": self.today_coin_cash,
            "yesterday_coin_cash": self.yesterday_coin_cash,
            "coin_balance": self.coin_balance,
            "cash_balance": round(cash_yuan, 2),
            "total_balance": round(total_yuan, 2),
            "coin_records": self.coin_records,
            "cash_records": self.cash_records,
            "ratio_param": SELECTED_PARAM
        }


def load_cookies_from_env():
    """从环境变量ksck加载cookie配置，格式：备注#cookie#salt#ip|端口|用户名|密码|到期日期"""
    cookies = []

    ksck_env = os.getenv('ksck')
    if not ksck_env:
        print("未找到环境变量 ksck")
        return cookies

    lines = ksck_env.strip().split('&')

    for i, line in enumerate(lines):
        line = line.strip()
        if not line:
            continue

        parts = line.split('#')
        if len(parts) < 3:
            print(f"第{i + 1}行格式错误，跳过: {line}")
            continue

        remark = parts[0].strip()
        cookie = parts[1].strip()
        salt = parts[2].strip()  # 暂时不使用

        if not cookie:
            print(f"第{i + 1}行cookie为空，跳过")
            continue

        # 不再检查socName，直接添加
        cookies.append({
            "cookie": cookie,
            "remark": remark
        })

    print(f"从环境变量 ksck 加载了 {len(cookies)} 个账号")
    return cookies


def create_excel_summary(data_list, filename="快手金币数据汇总.xlsx"):
    """创建Excel汇总表格"""
    wb = Workbook()
    ws = wb.active
    ws.title = "金币数据汇总"

    # 设置表头
    headers = ["序号", "用户名", "备注", "今日金币", "昨日金币", "当前金币",
               f"今日{SELECTED_PARAM}后(元)", f"昨日{SELECTED_PARAM}后(元)", "当前现金(元)", "累计收益(元)",
               "最近金币流水", "最近现金流水"]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        cell.fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")

    # 填充数据
    for row, data in enumerate(data_list, 2):
        ws.cell(row=row, column=1, value=data["index"]).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=data["user_name"])
        ws.cell(row=row, column=3, value=data["remark"])
        ws.cell(row=row, column=4, value=data["today_coin"]).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=5, value=data["yesterday_coin"]).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=6, value=data["coin_balance"]).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=7, value=data["today_coin_cash"]).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=8, value=data["yesterday_coin_cash"]).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=9, value=data["cash_balance"]).alignment = Alignment(horizontal="right")
        ws.cell(row=row, column=10, value=data["total_balance"]).alignment = Alignment(horizontal="right")

        # 格式化流水记录
        coin_flow = "\n".join([f"{r['time']} {r['title']} +{r['amount']}" for r in data["coin_records"]])
        cash_flow = "\n".join([f"{r['time']} {r['title']} {r['amount']}元" for r in data["cash_records"]])

        ws.cell(row=row, column=11, value=coin_flow)
        ws.cell(row=row, column=12, value=cash_flow)

    # 设置列宽
    column_widths = {
        'A': 8,  # 序号
        'B': 20,  # 用户名
        'C': 20,  # 备注
        'D': 12,  # 今日金币
        'E': 12,  # 昨日金币
        'F': 12,  # 当前金币
        'G': 18,  # 今日兑换后
        'H': 18,  # 昨日兑换后
        'I': 15,  # 当前现金
        'J': 15,  # 累计收益
        'K': 30,  # 最近金币流水
        'L': 30  # 最近现金流水
    }

    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # 设置行高
    for row in range(2, len(data_list) + 2):
        ws.row_dimensions[row].height = 60

    # 保存文件
    wb.save(filename)
    print(f"\nExcel表格已生成: {filename}")

    # 打印汇总统计
    total_today = sum(data["today_coin"] for data in data_list)
    total_yesterday = sum(data["yesterday_coin"] for data in data_list)
    total_today_cash = sum(data["today_coin_cash"] for data in data_list)
    total_yesterday_cash = sum(data["yesterday_coin_cash"] for data in data_list)
    total_coins = sum(data["coin_balance"] for data in data_list)
    total_cash = sum(data["cash_balance"] for data in data_list)
    total_accumulative = sum(data["total_balance"] for data in data_list)

    ratio = RATIO_MAP.get(SELECTED_PARAM, 0.00006)
    print(f"\n汇总统计 (使用比率参数 {SELECTED_PARAM}, 比率: {ratio}):")
    print(f"今日金币总数: {total_today}")
    print(f"今日金币兑换后: {total_today_cash}元")
    print(f"昨日金币总数: {total_yesterday}")
    print(f"昨日金币兑换后: {total_yesterday_cash}元")
    print(f"当前金币总数: {total_coins}")
    print(f"当前现金总数: {total_cash}元")
    print(f"累计收益总数: {total_accumulative}元")


def main():
    """主函数"""
    print_disclaimer()
    print("开始查询快手账号数据...")
    print("=" * 50)
    print(f"当前使用比率参数: {SELECTED_PARAM}, 兑换比率: {RATIO_MAP.get(SELECTED_PARAM, 0.00006)}")
    print("=" * 50)

    # 从环境变量加载cookie配置
    cookies_config = load_cookies_from_env()

    if not cookies_config:
        print("错误: 未找到有效的cookie配置")
        print("\n请设置环境变量 ksck，格式为:")
        print("备注#cookie#salt#ip|端口|用户名|密码|到期日期")
        print("每行一个账号，用&分隔，例如:")
        print("我的账号1#kuaishou.api_st=xxx; token=xxx#salt1#&我的账号2#kuaishou.api_st=yyy; token=yyy#salt2#")
        return

    all_data = []
    success_count = 0

    for i, account in enumerate(cookies_config):
        if not account.get("cookie"):
            print(f"账号[{i + 1}] cookie为空，跳过")
            continue

        client = KuaishouClient(
            index=i + 1,
            cookie=account["cookie"],
            remark=account["remark"]
        )

        if client.query_all_data():
            data = client.get_display_data()
            all_data.append(data)
            success_count += 1

            # 打印单个账号信息
            print(f"\n账号[{i + 1}] 用户名: {data['user_name']} 备注: {account['remark']} 查询成功:")
            print(f"  今日金币: {data['today_coin']} → {data['today_coin_cash']}元")
            print(f"  昨日金币: {data['yesterday_coin']} → {data['yesterday_coin_cash']}元")
            print(f"  当前金币: {data['coin_balance']}")
            print(f"  当前现金: {data['cash_balance']}元")
            print(f"  累计收益: {data['total_balance']}元")

            if data['coin_records']:
                print("  最近金币流水:")
                for record in data['coin_records']:
                    print(f"    {record['time']} {record['title']} +{record['amount']}")

            if data['cash_records']:
                print("  最近现金流水:")
                for record in data['cash_records']:
                    print(f"    {record['time']} {record['title']} {record['amount']}元")
        else:
            print(f"\n账号[{i + 1}] {account['remark']} 查询失败")

        # 添加0.5秒延迟，避免请求过于频繁
        if i < len(cookies_config) - 1:  # 最后一个账号不需要延迟
            print("等待0.5秒后继续下一个账号...")
            time.sleep(0.5)

    # 生成Excel表格
    if all_data:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"快手金币数据汇总_{SELECTED_PARAM}后_{timestamp}.xlsx"
        create_excel_summary(all_data, filename)
        print(f"\n成功查询 {success_count}/{len(cookies_config)} 个账号")
    else:
        print("没有成功查询到任何账号数据")


if __name__ == "__main__":
    main()